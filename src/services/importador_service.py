import sqlite3
import os

from openpyxl import load_workbook


class ImportadorService:

    def __init__(self):

        self.caminho_db = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "dados",
            "tributacao.db"
        )

    def importar_ncm(self, arquivo_excel):

        conn = sqlite3.connect(self.caminho_db)

        cursor = conn.cursor()

        arquivo_excel = os.path.abspath(arquivo_excel)

        print("Arquivo:", arquivo_excel)

        wb = load_workbook(arquivo_excel)

        ws = wb.active

        quantidade = 0

        for linha in ws.iter_rows(min_row=2, values_only=True):

            codigo = linha[0]
            descricao = linha[1]

            if not codigo:
                continue

            cursor.execute("""
                INSERT OR IGNORE INTO ncm
                (codigo, descricao)
                VALUES (?, ?)
            """, (str(codigo), descricao))

            quantidade += 1

        conn.commit()
        conn.close()

        return quantidade